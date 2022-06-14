# import necessary modules
from email.policy import strict
import os
import sys
import tkinter
from tkinter.messagebox import QUESTION
from matplotlib.pyplot import get
from openpyxl import *
from tkinter import *
from datetime import datetime
import simpleaudio as sa

# globally declare wb and sheet variable and opening the existing excel file
wb = load_workbook('./Excel Database 2.xlsx')
sheet = wb.active

# create global variables
idx = 0  #counter
start_time = None #timestamp
end_time = None #timestamp
emotion = None #widget
emotion_rating = None #widget
sound_list  = None #list of sound to be played per group selection
current_selection = None  #sound group selection from dropdown menu
file_path_dict = {"Animal Sound Block 1":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Animal/Block 1",  
				  "Animal Sound Block 2":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Animal/Block 2",
				  "Nature Sound Block 1":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Nature/Block 1", 
				  "Nature Sound Block 2":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Nature/Block 2",
				  "Nature Sound Block 3":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Nature/Block 3",  
				  "People Sound":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/People/Block 1",
				  "Transport Sound Block 1":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Transport/Block 1",
				  "Transport Sound Block 2":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Transport/Block 2",
				  "Practice Sound":"./IADS-E/IADS-E sound stimuli (IADS-2 is not included)/Practice sound"}


#create a function that retrieves the sound group&blocks from dropdown menu and returns the file_path using globally defined _file path dictionary
def compile_sound_group(folder):
	for key, value in file_path_dict.items():
		if folder == key:
			file_path = value
	else:
		pass
	return file_path

#Function that uses the file_path to compile list of all audio files within the input path folder
def compile_sounds(dir_path):
	extension = (".wav")
	sound = []
	for files in os.listdir(dir_path):
		if files.endswith(extension):
			sound.append(files)
		else:
			pass
	return sound

#Logic steps that takes when a dropmenu option is selected
def get_sound_group(sound_group_selection):
	global current_selection
	current_selection = sound_group_selection
	global sound_list, idx #redefine the global soundlist and counter
	path = compile_sound_group(sound_group_selection) #obtain the file_path
	sound_list = compile_sounds(path) #compile the sound using the file_path
	idx = 0 # reset the counter
	refresh_screen(idx) #refresh the screen

#Function for excel to handle excel dimension and headers
def excel():
	# resize the width of columns in excel spreadsheet
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 30
	sheet.column_dimensions['C'].width = 30
	sheet.column_dimensions['D'].width = 30
	sheet.column_dimensions['E'].width = 30
	sheet.column_dimensions['F'].width = 30
	sheet.column_dimensions['G'].width = 30
	sheet.column_dimensions['H'].width = 30

	# write given data to an excel spreadsheet at particular location
	sheet.cell(row=1, column=1).value = "SoundID"
	sheet.cell(row=1, column=2).value = "Starttime"
	sheet.cell(row=1, column=3).value = "Emotion"
	sheet.cell(row=1, column=4).value = "Emotionrating"
	sheet.cell(row=1, column=5).value = "Endtime"
	sheet.cell(row=1, column=6).value = "UserID"
	sheet.cell(row=1, column=7).value = "Soundgroup"
	sheet.cell(row=1, column=8).value = "Soundgroupselection"

# Functions to take data from GUI window and write to an excel file
def insert():
	global start_time
	global end_time
	global idx
	global emotion, emotion_rating
	end_time = get_end_time()

	# assigning the max row and max column value upto which data is written in an excel sheet to the variable
	current_row = sheet.max_row
	current_column = sheet.max_column

	# get method returns current text as string which we write into excel spreadsheet at particular location
	sheet.cell(row=current_row + 1, column=1).value = sound_list[idx].split(".")[0] #gets the string, splits it and select the first element
	sheet.cell(row=current_row + 1, column=2).value = start_time
	sheet.cell(row=current_row + 1, column=3).value = emotion.get()
	sheet.cell(row=current_row + 1, column=4).value = emotion_rating.get()
	sheet.cell(row=current_row + 1, column=5).value = end_time
	sheet.cell(row=current_row + 1, column=6).value = user_id_field.get()
	sheet.cell(row=current_row + 1, column=7).value = sound_group.get().split(" ")[0] #gets the string, splits it and select the first element
	sheet.cell(row=current_row + 1, column=8).value = sound_group.get()
	# save the file
	wb.save('./Excel Database 2.xlsx')

	#update the id and refresh screen or end program
	idx += 1
	if idx < len(sound_list):
		refresh_screen(idx)
	else:
		# inform user that category sounds has been completed
		tkinter.messagebox.showinfo(title = "Info", message = "End of this Category, Click next category") #display message box
		
	#Function to refresh screen
def refresh_screen(question_idx):
	global sound_list
	update_number = 1
	emotion_rating.set("How intense was the emotion?") # reset emotion intensity to default value
	emotion.set("Please select emotion experienced") # reset emotion dropdown to default value
	sound_label.config(text= f'Sound {question_idx + update_number} of {len(sound_list)}')  #update sound counter label
	return emotion_rating, emotion

	#Function to collect startime datetime
def get_start_time():
	current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
	return current_time

	#Function to playsound
def play():
	global file_path_dict
	global current_selection
	path = file_path_dict[current_selection] #path of sound group selection
	global start_time
	global sound_list
	global idx
	start_time = get_start_time() #get starttime upon clicking plays sound
	wave_obj = sa.WaveObject.from_wave_file(path + "/" + sound_list[idx]) #input combine path of soundfile
	play_obj = wave_obj.play() #play soundfile
	play_obj.wait_done()

	#Function to collect endtime timestamp
def get_end_time():
	current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
	return current_time

 
# Driver code
if __name__ == "__main__":

	root = Tk() # create a GUI window
	root.configure(background='light green') # set the background colour of GUI window
	root.title("Data Annotation Form")# set the title of GUI window
	root.geometry("925x400")# set the configuration of GUI window
	excel()

	#create textbox and label for userid and position
	user_id = Label(root,text="User ID",fg="Black", bg="Light Blue")
	user_id_field = Entry(root)
	user_id.grid(row=0, column=0,)
	user_id_field.grid(row=0,column=1,ipadx="30")
	place_holder = Label(root,text="").grid(row=1,column=0)

	#create sound group selection label and button
	sound_group_selection = Label(root, text="Select sound group",bg="Light Blue")
	sound_group_selection.grid(row=0, column=3)

	sound_group = StringVar(root)
	sound_group.set("Which sound group? ") # default value
	w3 = OptionMenu(root, sound_group, "Practice Sound","Nature Sound Block 1","Nature Sound Block 2",
										"Nature Sound Block 3", "Animal Sound Block 1","Animal Sound Block 2", 
										"People Sound", "Transport Sound Block 1","Transport Sound Block 2",
										command = get_sound_group)
	w3.grid(row=0, column=4, ipadx="30")

    #create sound play button and position
	play_button = Button(root, text="Play Sound", fg="Black", bg="Light Blue", command = lambda: play())
	play_button.grid(row=2, column=1,ipadx="40")

	#create Soundlabel and position
	sound_label = Label (root, text= f'Play sound {idx + 1} of 30' , fg="Black", bg="Light Blue")
	sound_label.grid(row=2, column=0)
	place_holder2 = Label(root,text="").grid(row=3,column=0)
	
	#create an emotion dropdown button and position
	emotion = StringVar(root)
	emotion.set("Please select emotion experienced") # default value
	w3 = OptionMenu(root, emotion, "Sadness", "Fear", "Happiness")
	w3.grid(row=4, column=1, ipadx="30")

	#Create emotion dropdown label
	emotion_label = Label (root,text= f'Select emotion' , fg="Black", bg="Light Blue")
	emotion_label.grid(row=4,column=0,)
	place_holder3 = Label(root,text="").grid(row=5,column=0)

	#create an emotion rating dropdown button and position
	emotion_rating = StringVar(root)
	emotion_rating.set("How intense was the emotion?") # default value
	w4 = OptionMenu(root, emotion_rating, "1", "2", "3", "4", "5", "6", "7", "8", "9")
	w4.grid(row=6, column=1, ipadx="30")

	#create emotion rating dropdown label
	emotion_rating_label = Label (root,text= 'Emotion Intensity' , fg="Black", bg="Light Blue")
	emotion_rating_label.grid(row=6,column=0)

	# call excel function
	excel()

	# create a Submit Button and place into the root window
	placeholder4 = Label(root, text="").grid(row=7,column=0)
	placeholder5 = Label(root, text="").grid(row=8,column=0)
	submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
	submit.grid(row=9, column=2)
	# start the GUI
	root.mainloop()